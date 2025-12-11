"""
导出与快速核对模块。
负责将处理后的数据导出为 Excel 文件（支持本地文件和内存流），并生成简要的数据核对报告。
"""
import re
from datetime import datetime
from io import BytesIO
from typing import Dict, List, Tuple, Union, Any

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.workbook import Workbook


def generate_excel_bytes(
    items: List[Dict[str, Any]], 
    product_url_from_file: str = "",
    base_name: str = "",
    strategy_name: str = ""
) -> Tuple[BytesIO, str]:
    """
    生成 Excel 文件的内存流和建议文件名。
    """
    file_name = _generate_filename(items, base_name, strategy_name)
    output = BytesIO()
    
    # 执行导出逻辑
    _write_excel_data(items, output, product_url_from_file, strategy_name)
    
    # 指针回到开头
    output.seek(0)
    
    return output, file_name


def export_to_excel(
    items: List[Dict[str, Any]],
    path: str = "output.xlsx",
    product_url_from_file: str = "",
    base_name: str = "",
    strategy_name: str = ""
) -> str:
    """
    将 items 导出为本地 Excel 文件。
    """
    # 处理路径逻辑
    if path == "output.xlsx" or path.endswith("/") or path.endswith("\\"):
        file_name = _generate_filename(items, base_name, strategy_name)
        if path.endswith("/") or path.endswith("\\"):
            path = os.path.join(path, file_name)
        else:
            path = file_name
            
    _write_excel_data(items, path, product_url_from_file, strategy_name)
    return path


def quick_check(items: List[Dict[str, Any]]) -> Dict[str, int]:
    """
    生成快速核对报告，统计关键指标。
    """
    total_skus = len(items)
    missing_price_skus = sum(1 for item in items if not item.get("price"))
    zero_cost_skus = sum(1 for item in items if item.get("cost") == 0.0)
    zero_suggested_price_skus = sum(1 for item in items if item.get("suggested_price") == 0.0)

    return {
        "总SKU数": total_skus,
        "价格缺失SKU数": missing_price_skus,
        "成本为0的SKU数": zero_cost_skus,
        "建议售价为0的SKU数": zero_suggested_price_skus,
    }


# ==========================================
# 内部辅助函数
# ==========================================

def _generate_filename(items: List[Dict[str, Any]], base_name: str = "", strategy_name: str = "") -> str:
    """根据商品信息或基础文件名生成文件名。"""
    file_name = None
    
    # 1. 如果提供了基础文件名 (来自导入)，则优先使用
    if base_name:
        # 去掉可能的扩展名
        base_name = os.path.splitext(base_name)[0]
        file_name = base_name
        
        # 附加策略名
        if strategy_name:
            # 简单的策略名映射
            strat_map = {
                "default": "默认定价", 
                "limited": "限时限量",
                "roi": "最佳投产计算",
                "equilibrium": "智能平衡定价"
            }
            suffix = strat_map.get(strategy_name, strategy_name)
            file_name = f"{file_name}_{suffix}"
            
    # 2. 否则尝试使用商品标题
    elif items and items[0].get("product_title_main"):
        file_name = items[0].get("product_title_main", "").strip()

    # 3. 尝试使用第一个 SKU 名称
    if not file_name:
        for item in items:
            if item.get("name"):
                file_name = item.get("name")
                break
    
    # 4. 默认文件名
    if not file_name:
        file_name = "商品数据"
        
    # 清理非法字符
    safe_filename = re.sub(r'[\\/:*?"<>|]', '_', file_name)
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    
    return f"{safe_filename}_{timestamp}.xlsx"


def _write_excel_data(
    items: List[Dict[str, Any]], 
    target: Union[str, BytesIO], 
    product_url: str,
    strategy_name: str = ""
):
    """核心导出逻辑：写入数据并调用格式化。"""
    rows = []
    
    # 确定商品链接
    if not product_url and items:
        product_url = items[0].get("product_url", "")

    # 动态确定定价列名 (保持与 UI 一致)
    price_label = "建议售价/活动价"
    if strategy_name == "limited":
        price_label = "限时限量购价格"
    elif strategy_name == "equilibrium":
        price_label = "智能平衡建议价"
    elif strategy_name == "default":
        price_label = "默认毛利建议价"
    # 兼容旧逻辑：如果策略名为空但数据里有特征key
    elif items and items[0].get("limited_time_price"):
        price_label = "限时限量购价格"

    for sku_data in items:
        row = {
            "SKU名": sku_data.get("name", ""),
            "价格(元)": sku_data.get("price"),
            "库存(件数)": sku_data.get("stock", ""),
            price_label: sku_data.get("suggested_price", ""),
            "实际成交价": sku_data.get("selling_price", ""),
            "运费(元)": sku_data.get("overall_shipping_cost", ""),
            "保本投产": sku_data.get("breakeven_roi", ""),
            "净投产": sku_data.get("net_roi", ""),
            "★最佳投产比": sku_data.get("best_roi", ""),
            "利润/单": sku_data.get("profit_per_order", ""),
            
            # 智能平衡定价策略专有
            "保本底价": sku_data.get("breakeven_price", ""),
            "预期广告费占比": f"{1/sku_data.get('expected_roi'):.1%}" if sku_data.get("expected_roi") else "",
            "广告费上限": sku_data.get("ad_cost_limit", "")
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    df.to_excel(target, index=False, sheet_name="商品数据")
    
    _format_excel(target, product_url)


def _format_excel(target: Union[str, BytesIO], product_url: str):
    """对 Excel 文件进行美化格式化。"""
    wb = load_workbook(target)
    ws = wb.active

    # 写入商品链接
    ws['K1'] = product_url
    ws['K1'].font = Font(underline="single", color="0000FF")
    ws['K1'].hyperlink = product_url

    # 样式定义
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # 格式化表头
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = border

    # 格式化数据行
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.border = border
            cell.alignment = left_align

    # 自动调整列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except Exception:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    ws.freeze_panes = "A2"
    wb.save(target)
